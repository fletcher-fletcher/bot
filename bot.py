import logging
import sqlite3
import asyncio
import os
import csv
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Optional

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, BufferedInputFile
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode

# ==================== НАСТРОЙКИ ====================
BOT_TOKEN = "8781874817:AAHLeiKjpLEe41ADa3NMUQCqqcTfitQZV2c"  
ADMIN_IDS = [628687487, 5853079155]  

# ==================== ИНИЦИАЛИЗАЦИЯ ====================
logging.basicConfig(level=logging.INFO)

# Создаем бота и диспетчер для aiogram 3.x
bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
storage = MemoryStorage()
dp = Dispatcher(storage=storage)

# Варианты для поля "Кто вы"
PROFESSION_OPTIONS = [
    "Предприниматель",
    "Юрист", 
    "Бухгалтер",
    "Наёмный сотрудник",
    "Другое"
]

# ==================== БАЗА ДАННЫХ ====================
def init_db():
    """Создание таблиц при первом запуске"""
    conn = sqlite3.connect('efir_bot.db')
    cur = conn.cursor()
    
    # Таблица эфиров
    cur.execute('''
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            title TEXT NOT NULL,
            room_link TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Таблица регистраций (расширенная)
    cur.execute('''
        CREATE TABLE IF NOT EXISTS registrations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            event_id INTEGER NOT NULL,
            username TEXT,
            full_name TEXT NOT NULL,
            phone TEXT NOT NULL,
            profession TEXT NOT NULL,
            registered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, event_id)
        )
    ''')
    
    conn.commit()
    conn.close()

# ==================== ФУНКЦИИ РАБОТЫ С БД ====================
def create_event(code: str, title: str, room_link: str) -> bool:
    """Создать новый эфир"""
    try:
        conn = sqlite3.connect('efir_bot.db')
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO events (code, title, room_link) VALUES (?, ?, ?)",
            (code, title, room_link)
        )
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()

def get_event_by_code(code: str) -> Optional[dict]:
    """Получить информацию об эфире по его коду"""
    conn = sqlite3.connect('efir_bot.db')
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM events WHERE code = ?", (code,))
    event = cur.fetchone()
    conn.close()
    return dict(event) if event else None

def get_event_by_id(event_id: int) -> Optional[dict]:
    """Получить информацию об эфире по ID"""
    conn = sqlite3.connect('efir_bot.db')
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM events WHERE id = ?", (event_id,))
    event = cur.fetchone()
    conn.close()
    return dict(event) if event else None

def check_registration(user_id: int, event_id: int) -> bool:
    """Проверить, регистрировался ли пользователь на этот эфир"""
    conn = sqlite3.connect('efir_bot.db')
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM registrations WHERE user_id = ? AND event_id = ?",
        (user_id, event_id)
    )
    result = cur.fetchone()
    conn.close()
    return result is not None

def save_registration(user_id: int, event_id: int, username: str, full_name: str, phone: str, profession: str) -> bool:
    """Сохранить регистрацию пользователя"""
    try:
        conn = sqlite3.connect('efir_bot.db')
        cur = conn.cursor()
        cur.execute(
            """INSERT INTO registrations 
               (user_id, event_id, username, full_name, phone, profession) 
               VALUES (?, ?, ?, ?, ?, ?)""",
            (user_id, event_id, username, full_name, phone, profession)
        )
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()

def get_registrations_count(event_id: int) -> int:
    """Получить количество регистраций на эфир"""
    conn = sqlite3.connect('efir_bot.db')
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM registrations WHERE event_id = ?", (event_id,))
    count = cur.fetchone()[0]
    conn.close()
    return count

def export_event_registrations(event_code: str):
    """Экспорт регистраций на эфир (для админа)"""
    conn = sqlite3.connect('efir_bot.db')
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT r.*, e.title as event_title 
        FROM registrations r
        JOIN events e ON r.event_id = e.id
        WHERE e.code = ?
        ORDER BY r.registered_at
    """, (event_code,))
    registrations = cur.fetchall()
    conn.close()
    return registrations

# ==================== СОСТОЯНИЯ FSM ====================
class Registration(StatesGroup):
    waiting_for_full_name = State()
    waiting_for_phone = State()
    waiting_for_profession = State()
    waiting_for_custom_profession = State()

# ==================== КОМАНДЫ АДМИНА ====================
@dp.message(Command("new"))
async def cmd_new_event(message: types.Message):
    """Создание нового эфира (только для админа)"""
    if message.from_user.id not in ADMIN_IDS:
        await message.reply("⛔ У вас нет прав на выполнение этой команды.")
        return
    
    try:
        # Парсим команду в формате: /new Код | Название | Ссылка
        # Пример: /new may2025 | Майский эфир 2025 | https://zoom.us/j/123
        
        # В aiogram 3.x текст команды получаем так
        command_parts = message.text.split(maxsplit=1)
        if len(command_parts) < 2:
            await message.reply(
                "❌ Неправильный формат. Используйте:\n"
                "/new КОД | НАЗВАНИЕ | ССЫЛКА НА КОМНАТУ\n\n"
                "Пример: /new may2025 | Майский эфир 2025 | https://zoom.us/j/123"
            )
            return
        
        parts = command_parts[1].strip().split('|')
        if len(parts) < 3:
            await message.reply(
                "❌ Неправильный формат. Используйте:\n"
                "/new КОД | НАЗВАНИЕ | ССЫЛКА НА КОМНАТУ\n\n"
                "Пример: /new may2025 | Майский эфир 2025 | https://zoom.us/j/123"
            )
            return
        
        code = parts[0].strip()
        title = parts[1].strip()
        room_link = parts[2].strip()
        
        # Проверяем ссылку (хотя бы базово)
        if not room_link.startswith(('http://', 'https://')):
            await message.reply("❌ Ссылка должна начинаться с http:// или https://")
            return
        
        if create_event(code, title, room_link):
            # Создаем ссылку для поста
            bot_info = await bot.me()
            bot_link = f"https://t.me/{bot_info.username}?start={code}"
            
            response = (
                f"✅ Эфир успешно создан!\n\n"
                f"📌 Код: {code}\n"
                f"📝 Название: {title}\n"
                f"🔗 Комната: {room_link}\n\n"
                f"🔗 Ссылка для поста в канале:\n"
                f"<code>{bot_link}</code>\n\n"
                f"📊 Статистика будет доступна по команде:\n"
                f"/stats {code}\n"
                f"📥 CSV: /csv {code}\n"
                f"📥 Excel: /xls {code}"
            )
            
            await message.reply(response)
        else:
            await message.reply("❌ Эфир с таким кодом уже существует!")
            
    except Exception as e:
        await message.reply(f"❌ Ошибка: {str(e)}")

@dp.message(Command("stats"))
async def cmd_event_stats(message: types.Message):
    """Статистика по конкретному эфиру"""
    if message.from_user.id not in ADMIN_IDS:
        await message.reply("⛔ У вас нет прав на выполнение этой команды.")
        return
    
    # В aiogram 3.x аргументы получаем так
    command_parts = message.text.split()
    if len(command_parts) < 2:
        await message.reply("❌ Укажите код эфира. Пример: /stats may2025")
        return
    
    args = command_parts[1]
    
    registrations = export_event_registrations(args)
    
    if not registrations:
        await message.reply(f"📭 На эфир с кодом '{args}' пока никто не зарегистрировался")
        return
    
    # Формируем статистику
    event_title = registrations[0]['event_title']
    response = f"📊 Статистика по эфиру: {event_title}\n"
    response += f"📌 Код: {args}\n"
    response += f"👥 Всего регистраций: {len(registrations)}\n\n"
    response += "📋 Список участников:\n"
    
    for i, reg in enumerate(registrations, 1):
        response += f"{i}. {reg['full_name']}\n"
        response += f"   📱 {reg['phone']}\n"
        response += f"   💼 {reg['profession']}\n"
        response += f"   🆔 @{reg['username'] if reg['username'] else 'нет'}\n"
        response += f"   🕐 {reg['registered_at'][:16]}\n\n"
        
        # Telegram не любит слишком длинные сообщения
        if len(response) > 3500:
            response += "... (продолжение в следующем сообщении)"
            await message.reply(response)
            response = ""
    
    if response:
        await message.reply(response)

@dp.message(Command("events"))
async def cmd_list_events(message: types.Message):
    """Список всех эфиров"""
    if message.from_user.id not in ADMIN_IDS:
        await message.reply("⛔ У вас нет прав на выполнение этой команды.")
        return
    
    conn = sqlite3.connect('efir_bot.db')
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM events ORDER BY created_at DESC")
    events = cur.fetchall()
    conn.close()
    
    if not events:
        await message.reply("📭 Пока нет созданных эфиров")
        return
    
    response = "📋 Все эфиры:\n\n"
    for event in events:
        count = get_registrations_count(event['id'])
        response += (
            f"🔹 {event['title']}\n"
            f"   Код: {event['code']}\n"
            f"   Участников: {count}\n"
            f"   Создан: {event['created_at'][:16]}\n"
            f"   /stats {event['code']}\n"
            f"   /csv {event['code']}\n"
            f"   /xls {event['code']}\n\n"
        )
    
    await message.reply(response)

# ==================== ЭКСПОРТ В CSV ====================
@dp.message(Command("csv"))
async def cmd_export_csv(message: types.Message):
    """Экспорт регистраций в CSV (только для админа)"""
    if message.from_user.id not in ADMIN_IDS:
        await message.reply("⛔ У вас нет прав на выполнение этой команды.")
        return
    
    # Получаем код эфира
    command_parts = message.text.split()
    if len(command_parts) < 2:
        await message.reply("❌ Укажите код эфира. Пример: /csv may2025")
        return
    
    event_code = command_parts[1]
    
    # Получаем данные
    registrations = export_event_registrations(event_code)
    
    if not registrations:
        await message.reply(f"📭 На эфир с кодом '{event_code}' никто не зарегистрировался")
        return
    
    # Получаем название эфира
    event_title = registrations[0]['event_title']
    
    # СОЗДАЕМ CSV ФАЙЛ
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Заголовки
    writer.writerow(['Имя', 'Телефон', 'Профессия', 'Telegram Username', 'Дата регистрации', 'ID пользователя'])
    
    # Данные
    for reg in registrations:
        writer.writerow([
            reg['full_name'],
            reg['phone'],
            reg['profession'],
            f"@{reg['username']}" if reg['username'] else '-',
            reg['registered_at'][:16],
            reg['user_id']
        ])
    
    # Получаем байты CSV
    csv_bytes = output.getvalue().encode('utf-8-sig')
    
    # СОЗДАЕМ ФАЙЛ ДЛЯ ОТПРАВКИ
    file = BufferedInputFile(
        file=csv_bytes,
        filename=f"registrations_{event_code}.csv"
    )
    
    # ОТПРАВЛЯЕМ ФАЙЛ
    await message.reply_document(
        document=file,
        caption=f"📊 CSV-экспорт по эфиру:\n{event_title}\n"
                f"📌 Код: {event_code}\n"
                f"👥 Всего участников: {len(registrations)}"
    )

# ==================== ЭКСПОРТ В EXCEL ====================
@dp.message(Command("xls"))
async def cmd_export_xls(message: types.Message):
    """Экспорт регистраций в Excel (только для админа)"""
    if message.from_user.id not in ADMIN_IDS:
        await message.reply("⛔ У вас нет прав на выполнение этой команды.")
        return
    
    # Получаем код эфира
    command_parts = message.text.split()
    if len(command_parts) < 2:
        await message.reply("❌ Укажите код эфира. Пример: /xls may2025")
        return
    
    event_code = command_parts[1]
    
    # Получаем данные
    registrations = export_event_registrations(event_code)
    
    if not registrations:
        await message.reply(f"📭 На эфир с кодом '{event_code}' никто не зарегистрировался")
        return
    
    try:
        # Создаем Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Эфир {event_code}"
        
        # Заголовки
        headers = ['№', 'Имя', 'Телефон', 'Профессия', 'Telegram', 'Дата регистрации', 'ID пользователя']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Данные
        for row, reg in enumerate(registrations, 2):
            ws.cell(row=row, column=1, value=row-1).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=reg['full_name'])
            ws.cell(row=row, column=3, value=reg['phone'])
            ws.cell(row=row, column=4, value=reg['profession'])
            ws.cell(row=row, column=5, value=f"@{reg['username']}" if reg['username'] else "-")
            ws.cell(row=row, column=6, value=reg['registered_at'][:16])
            ws.cell(row=row, column=7, value=reg['user_id'])
        
        # Автоподбор ширины колонок
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, len(registrations) + 2):
                cell_value = ws[f"{column_letter}{row}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Добавляем границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, len(registrations) + 2):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # Сохраняем в память
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        # Получаем название эфира
        event_title = registrations[0]['event_title']
        
        # Создаем файл для отправки
        file = BufferedInputFile(
            file=excel_bytes.getvalue(),
            filename=f"registrations_{event_code}.xlsx"
        )
        
        # Отправляем файл
        await message.reply_document(
            document=file,
            caption=f"📊 Excel-отчет по эфиру:\n{event_title}\n"
                    f"📌 Код: {event_code}\n"
                    f"👥 Всего участников: {len(registrations)}"
        )
        
    except ImportError:
        await message.reply("❌ Библиотека openpyxl не установлена. Установите: pip install openpyxl")
    except Exception as e:
        await message.reply(f"❌ Ошибка при создании Excel: {e}")

# ==================== РЕГИСТРАЦИЯ ПОЛЬЗОВАТЕЛЕЙ ====================
@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    """Обработка команды /start с параметром или без"""
    
    # В aiogram 3.x аргументы получаем так
    command_parts = message.text.split()
    args = command_parts[1] if len(command_parts) > 1 else ""
    
    if not args:
        await message.reply(
            "👋 Добро пожаловать!\n\n"
            "Это бот для регистрации на прямые эфиры.\n"
            "Чтобы зарегистрироваться, перейдите по специальной ссылке из поста в канале."
        )
        return
    
    # Ищем эфир по коду
    event = get_event_by_code(args)
    if not event:
        await message.reply("❌ Эфир не найден или ссылка устарела.")
        return
    
    # Сохраняем ID эфира в состояние
    await state.update_data(event_id=event['id'], event_code=args)
    
    # Проверяем, не регистрировался ли уже
    if check_registration(message.from_user.id, event['id']):
        event = get_event_by_id(event['id'])
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔗 Перейти в комнату", url=event['room_link'])]
        ])
        await message.reply(
            f"🔔 Вы уже зарегистрированы на этот эфир!\n\n"
            f"🎥 {event['title']}\n\n"
            f"🔗 Ссылка для входа:",
            reply_markup=keyboard
        )
        return
    
    # Начинаем регистрацию
    await message.reply(
        f"📝 <b>Регистрация на эфир:</b>\n"
        f"<i>{event['title']}</i>\n\n"
        f"Пожалуйста, введите ваше <b>полное имя</b> (ФИО):"
    )
    await state.set_state(Registration.waiting_for_full_name)

@dp.message(Registration.waiting_for_full_name)
async def process_full_name(message: types.Message, state: FSMContext):
    """Получаем ФИО пользователя"""
    full_name = message.text.strip()
    if len(full_name.split()) < 2:  # Хотя бы имя и фамилия
        await message.reply("❌ Пожалуйста, введите полное имя (имя и фамилию):")
        return
    
    await state.update_data(full_name=full_name)
    await message.reply(
        "📞 Теперь введите ваш <b>номер телефона</b>:\n"
        "Например: +7 (999) 123-45-67"
    )
    await state.set_state(Registration.waiting_for_phone)

@dp.message(Registration.waiting_for_phone)
async def process_phone(message: types.Message, state: FSMContext):
    """Получаем телефон"""
    phone = message.text.strip()
    
    # Простая проверка (можно усложнить под конкретные форматы)
    if len(phone) < 10:
        await message.reply("❌ Слишком короткий номер. Введите корректный телефон:")
        return
    
    await state.update_data(phone=phone)
    
    # Создаем клавиатуру с вариантами профессий
    keyboard_builder = ReplyKeyboardBuilder()
    for prof in PROFESSION_OPTIONS:
        keyboard_builder.button(text=prof)
    keyboard_builder.adjust(2)
    
    await message.reply(
        "💼 Кто вы по роду деятельности?\n"
        "Выберите из списка или напишите свой вариант:",
        reply_markup=keyboard_builder.as_markup(resize_keyboard=True, one_time_keyboard=True)
    )
    await state.set_state(Registration.waiting_for_profession)

@dp.message(Registration.waiting_for_profession)
async def process_profession(message: types.Message, state: FSMContext):
    """Обрабатываем выбор профессии"""
    profession = message.text.strip()
    
    if profession == "Другое":
        await message.reply(
            "✍️ Напишите ваш вариант:",
            reply_markup=ReplyKeyboardRemove()
        )
        await state.set_state(Registration.waiting_for_custom_profession)
    else:
        await complete_registration(message, state, profession)

@dp.message(Registration.waiting_for_custom_profession)
async def process_custom_profession(message: types.Message, state: FSMContext):
    """Обрабатываем свой вариант профессии"""
    profession = message.text.strip()
    if len(profession) < 2:
        await message.reply("❌ Слишком короткое значение. Опишите подробнее:")
        return
    
    await complete_registration(message, state, profession)

async def complete_registration(message: types.Message, state: FSMContext, profession: str):
    """Завершаем регистрацию"""
    
    # Получаем все данные
    data = await state.get_data()
    event = get_event_by_id(data['event_id'])
    
    if not event:
        await message.reply(
            "❌ Произошла ошибка. Эфир не найден.",
            reply_markup=ReplyKeyboardRemove()
        )
        await state.clear()
        return
    
    # Сохраняем регистрацию
    username = message.from_user.username or ""
    
    if save_registration(
        message.from_user.id,
        event['id'],
        username,
        data['full_name'],
        data['phone'],
        profession
    ):
        # Отправляем ссылку на комнату
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔗 Перейти в комнату", url=event['room_link'])]
        ])
        
        response = (
            f"✅ <b>Регистрация завершена!</b>\n\n"
            f"Спасибо, {data['full_name']}!\n"
            f"Вы зарегистрированы на эфир:\n"
            f"<i>{event['title']}</i>\n\n"
            f"🔗 <b>Ссылка для входа:</b>"
        )
        
        # Отправляем сообщение с клавиатурой
        await message.reply(
            response,
            reply_markup=keyboard
        )
        
        # Уведомляем админа о новой регистрации
        reg_count = get_registrations_count(event['id'])
        for admin_id in ADMIN_IDS:
            try:
                await bot.send_message(
                    admin_id,
                    f"📝 <b>Новая регистрация!</b>\n"
                    f"🎥 Эфир: {event['title']}\n"
                    f"👤 Имя: {data['full_name']}\n"
                    f"📞 Телефон: {data['phone']}\n"
                    f"💼 Кто: {profession}\n"
                    f"🆔 @{username if username else 'нет username'}\n"
                    f"📊 Всего на эфире: {reg_count}\n"
                    f"📥 CSV: /csv {data['event_code']}\n"
                    f"📥 Excel: /xls {data['event_code']}"
                )
            except:
                pass
    else:
        await message.reply(
            "❌ Ошибка при сохранении. Возможно, вы уже регистрировались на этот эфир.",
            reply_markup=ReplyKeyboardRemove()
        )
    
    await state.clear()

# ==================== КОМАНДА ОТМЕНЫ ====================
@dp.message(Command("cancel"))
async def cmd_cancel(message: types.Message, state: FSMContext):
    """Отмена текущего действия (работает всегда)"""
    current_state = await state.get_state()
    
    if current_state is None:
        await message.reply("❌ Нет активной регистрации для отмены.")
        return
    
    await state.clear()
    await message.reply(
        "✅ Регистрация отменена.",
        reply_markup=ReplyKeyboardRemove()
    )

# ==================== ЗАПУСК ====================
async def main():
    """Главная функция запуска бота"""
    init_db()
    print("="*50)
    print("🤖 Бот для регистрации на эфиры запущен!")
    print("="*50)
    print("\n📋 Команды администратора:")
    print("/new КОД | НАЗВАНИЕ | ССЫЛКА - создать эфир")
    print("/events - список всех эфиров")
    print("/stats КОД - статистика по эфиру")
    print("/csv КОД - выгрузить в CSV")
    print("/xls КОД - выгрузить в Excel")
    print("\n👤 Команды пользователей:")
    print("/start - начать работу с ботом")
    print("/cancel - отменить регистрацию")
    print("="*50)
    
    await dp.start_polling(bot)

if __name__ == '__main__':
    asyncio.run(main())
